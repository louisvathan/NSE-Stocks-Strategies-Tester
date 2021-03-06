import pandas as pd
import openpyxl as xl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Border, Side, Alignment, Font
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
plt.style.use('fivethirtyeight')


def rsi_plots(stock, new_df, plot_full_location):
    plt.figure(figsize=(12.2, 4.5))
    plt.title('RSI Plot ({})'.format(stock))
    plt.plot(new_df.index, new_df['RSI'])
    plt.axhline(0, linestyle='--', alpha = 0.5, color='gray')
    plt.axhline(10, linestyle='--', alpha = 0.5, color='orange')
    plt.axhline(20, linestyle='--', alpha = 0.5, color='green')
    plt.axhline(30, linestyle='--', alpha = 0.5, color='red')
    plt.axhline(70, linestyle='--', alpha = 0.5, color='red')
    plt.axhline(80, linestyle='--', alpha = 0.5, color='green')
    plt.axhline(90, linestyle='--', alpha = 0.5, color='red')
    plt.axhline(100, linestyle='--', alpha = 0.5, color='gray')
    plt.savefig('{}/{}_RSI.png'.format(plot_full_location, stock), dpi=150)
    rsi_hlink = '=HYPERLINK("{}\{}_RSI.png","RSI")'.format(plot_full_location, stock)
    plt.close()
    return rsi_hlink


def stock_sheet(stock, df, new_df, wb_rsi, left_alignment, center_alignment, 
                right_alignment, font, border, fmt_flo, stkplt_hlink, plot_full_location):
    #Transfer data from dataframes to Individual Stock Excel sheet.
    ws = wb_rsi.create_sheet(stock)
    
    #Writing Titles in Stock Sheet.
    ws.cell(1, 1, 'Date')
    ws.cell(1, 2, 'Open')
    ws.cell(1, 3, 'High')
    ws.cell(1, 4, 'Low')
    ws.cell(1, 5, 'Close')
    ws.cell(1, 6, 'Adj Close')
    ws.cell(1, 7, 'Volume')
    ws.cell(1, 8, 'RSI')
    for col in range(1, 9):
        cell = ws.cell(1, col)
        cell.font = font
        cell.border = border
        cell.alignment = center_alignment
    
    #Transferring data from dataframes to stock Excel sheet.
    for i in range(len(df)):    #From df
        cell = ws.cell(i+2, 1, str(df.loc[i, 'Date']))
        cell.alignment = left_alignment

        cell = ws.cell(i+2, 2, float(df.loc[i, 'Open']))
        cell.alignment = right_alignment
        cell.number_format = fmt_flo

        cell = ws.cell(i+2, 3, float(df.loc[i, 'High']))
        cell.alignment = right_alignment
        cell.number_format = fmt_flo

        cell = ws.cell(i+2, 4, float(df.loc[i, 'Low']))
        cell.alignment = right_alignment
        cell.number_format = fmt_flo

        cell = ws.cell(i+2, 5, float(df.loc[i, 'Close']))
        cell.alignment = right_alignment
        cell.number_format = fmt_flo

        cell = ws.cell(i+2, 6, float(df.loc[i, 'Adj Close']))
        cell.alignment = right_alignment
        cell.number_format = fmt_flo

        cell = ws.cell(i+2, 7, float(df.loc[i, 'Volume']))
        cell.alignment = right_alignment
        cell.number_format = fmt_flo
    
    for i in range(len(new_df)):    #From df that contains RSI Values.
        cell = ws.cell(i+2, 8, float(new_df.loc[i, 'RSI']))
        cell.number_format = fmt_flo
        cell.alignment = right_alignment
    
    cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    for row in range(1, len(df)+2):
        for col in cols:
            cell = ws['{}{}'.format(col, row)]
            cell.font = font
            cell.border = border
    
    #Adding HyperLinks for plots.
    #1.Stock Close Price Plot.
    cell = ws.cell(2, 11, stkplt_hlink)
    cell.alignment = left_alignment
    cell.font = font
    cell.border = border
    #2.Stock RSI Plot.
    rsi_hlink = rsi_plots(stock, new_df, plot_full_location)
    cell = ws.cell(2, 12, rsi_hlink)
    cell.alignment = left_alignment
    cell.font = font
    cell.border = border
    
    #Updating Cell Widths.
    dim_holder = DimensionHolder(worksheet=ws)
    for col in range(ws.min_column, ws.max_column+1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=12)
    ws.column_dimensions = dim_holder
    
    #Adding Color Coding for RSI Column.
    ws.conditional_formatting.add('H2:H{}'.format(len(new_df)+1),
                                  ColorScaleRule(start_type='percentile', start_value=10, start_color='00AA00',
                                                 mid_type='percentile', mid_value=50, mid_color='d4d4d6',
                                                 end_type='percentile', end_value=90, end_color='AA0000'))
    
    return ws


def rsi_summary(stock, df, new_df, wb_rsi, left_alignment, center_alignment, 
                right_alignment, font, border, fmt_flo, main_row, last_entry):
    #Creating a summary RSI sheet with all Stock values.
    ws_main = wb_rsi.active
    #Adding Titles
    ws_main.cell(1, 1, 'Stocks')
    ws_main.cell(1, 2, 'Cur. RSI Value')
    ws_main.cell(1, 3, 'Remarks')
    for col in range(1, 4):
        cell = ws_main.cell(1, col)
        cell.font = font
        cell.border = border
        cell.alignment = center_alignment
    
    #Adding Stock name.
    cell = ws_main.cell(main_row, 1, stock)
    cell.alignment = left_alignment
    cell.font = font
    cell.border = border
    
    #Adding Current RSI Value.
    cell = ws_main.cell(main_row, 2, float(new_df.loc[(len(new_df)-1), 'RSI']))
    cell.alignment = center_alignment
    cell.number_format = fmt_flo
    cell.font = font
    cell.border = border
    
    #Filling our Remarks Column based on last RSI Value.
    last_rsi = float(new_df.loc[(len(new_df)-1), 'RSI'])
    if last_rsi < 10:
        cell = ws_main.cell(main_row, 3, 'Very Strong Buy Signal')
    elif last_rsi > 10 and last_rsi < 20:
        cell = ws_main.cell(main_row, 3, 'Strong Buy Signal')
    elif last_rsi > 20 and last_rsi < 30:
        cell = ws_main.cell(main_row, 3, 'Buy Signal')
    elif last_rsi > 30 and last_rsi < 70:
        cell = ws_main.cell(main_row, 3, 'Neutral')
    elif last_rsi > 70 and last_rsi < 80:
        cell = ws_main.cell(main_row, 3, 'Sell Signal')
    elif last_rsi > 80 and last_rsi < 90:
        cell = ws_main.cell(main_row, 3, 'Strong Sell Signal')
    elif last_rsi > 90 and last_rsi < 100:
        cell = ws_main.cell(main_row, 3, 'Very Strong Sell Signal')
    cell.alignment = left_alignment
    cell.font = font
    cell.border = border

    #Adding Color Coding for Remarks column.
    red_text = Font(color="9C0006")
    red_fill = PatternFill(bgColor="FF8593")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", text="Very Strong Sell Signal", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("highlight",A1)))']
    ws_main.conditional_formatting.add('C2:C100', rule)

    red_text = Font(color="9C0006")
    red_fill = PatternFill(bgColor="FFA1AC")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", text="Strong Sell Signal", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("highlight",A1)))']
    ws_main.conditional_formatting.add('C2:C100', rule)

    red_text = Font(color="9C0006")
    red_fill = PatternFill(bgColor="FFC7CE")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", text="Sell Signal", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("highlight",A1)))']
    ws_main.conditional_formatting.add('C2:C100', rule)

    black_text = Font(color="000000")
    grey_fill = PatternFill(bgColor="B3B3B3")
    dxf = DifferentialStyle(font=black_text, fill=grey_fill)
    rule = Rule(type="containsText", operator="containsText", text="Neutral", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("highlight",A1)))']
    ws_main.conditional_formatting.add('C2:C100', rule)

    green_text = Font(color="008000")
    green_fill = PatternFill(bgColor="A1FFBA")
    dxf = DifferentialStyle(font=green_text, fill=green_fill)
    rule = Rule(type="containsText", operator="containsText", text="Buy Signal", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("highlight",A1)))']
    ws_main.conditional_formatting.add('C2:C100', rule)

    green_text = Font(color="008000")
    green_fill = PatternFill(bgColor="73FF98")
    dxf = DifferentialStyle(font=green_text, fill=green_fill)
    rule = Rule(type="containsText", operator="containsText", text="Strong Buy Signal", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("highlight",A1)))']
    ws_main.conditional_formatting.add('C2:C100', rule)

    green_text = Font(color="008000")
    green_fill = PatternFill(bgColor="63BE7B")
    dxf = DifferentialStyle(font=green_text, fill=green_fill)
    rule = Rule(type="containsText", operator="containsText", text="Very Strong Buy Signal", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("highlight",A1)))']
    ws_main.conditional_formatting.add('C2:C100', rule)
    
    #Updating First Sheet Name.
    if last_entry:
        ws_main.title = 'RSI Summary'
        last_entry = False
    
    #Updating Cell Widths.
    dim_holder = DimensionHolder(worksheet=ws_main)
    for col in range(ws_main.min_column, ws_main.max_column+1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws_main, min=col, max=col, width=12)
    ws_main.column_dimensions = dim_holder


def RSI(stock, df, wb_rsi, left_alignment, center_alignment, 
        right_alignment, font, border, fmt_flo, stkplt_hlink,
        plot_full_location, main_row, last_entry):
    #Intake dataframe and converts into RSI files(new_df).
    delta = df['Adj Close'].diff(1)
    delta = delta.dropna()

    up = delta.copy()
    down = delta.copy()

    up[up<0] = 0
    down[down>0] = 0

    period = 14
    avg_gain = up.rolling(window=period).mean()
    avg_loss = abs(down.rolling(window=period).mean())

    rs = avg_gain/avg_loss
    rsi = 100.0 - (100.0 / (1.0 + rs))

    new_df = pd.DataFrame()
    new_df['Adj Close'] = df['Adj Close']
    new_df['RSI'] = rsi
    
    ws = stock_sheet(stock, df, new_df, wb_rsi, left_alignment, center_alignment, 
                        right_alignment, font, border, fmt_flo, stkplt_hlink, plot_full_location)
    
    rsi_summary(stock, df, new_df, wb_rsi, left_alignment, center_alignment, 
                        right_alignment, font, border, fmt_flo, main_row, last_entry)
