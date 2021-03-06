import pandas as pd
import numpy as np
import openpyxl as xl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Border, Side, Alignment, Font
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
plt.style.use('fivethirtyeight')


def plot_macd(stock, plt_df, MACD, plot_full_location):
    plt_df = plt_df.set_index(pd.DatetimeIndex(plt_df['Date'].values))
    
    #Plotting MACD chart.
    plt.figure(figsize=(12.2, 4.5))
    plt.plot(plt_df.index, MACD, label = '{} MACD'.format(stock), color = 'red')
    plt.plot(plt_df.index, plt_df['Signal Line'], label = 'Signal Line', color = 'blue')
    plt.legend(loc='upper left')
    plt.savefig('{}/{}_MACD.png'.format(plot_full_location, stock), dpi=150)
    macd_hlink = '=HYPERLINK("{}\{}_MACD.png","MACD")'.format(plot_full_location, stock)
    plt.close()
    
    #Plotting Buy & Sell Signals.
    plt.figure(figsize=(12.2, 4.5))
    plt.scatter(plt_df.index, plt_df['Buy_Signal_Price'], color='green', label='Buy', marker='^', alpha=1)
    plt.scatter(plt_df.index, plt_df['Sell_Signal_Price'], color='red', label='Sell', marker='v', alpha=1)
    plt.plot(plt_df['Close'], label='Close Price', alpha=0.35)
    plt.title('Close Price Buy & Sell Signals ({})'.format(stock))
    plt.xlabel('Date')
    plt.ylabel('Close Price in INR')
    plt.legend(loc='upper left')
    plt.savefig('{}/{}_MACD_Signal.png'.format(plot_full_location, stock), dpi=150)
    signal_hlink = '=HYPERLINK("{}\{}_MACD_Signal.png","MACD SIGNAL")'.format(plot_full_location, stock)
    plt.close()
    
    return(macd_hlink, signal_hlink)
    
    
def stock_sheet(stock, df, wb_macd, left_alignment, center_alignment, 
                right_alignment, font, border, fmt_flo, stkplt_hlink, 
                plot_full_location, main_row, last_entry, hlinks):
    #Transfer data from dataframes to Individual Stock Excel sheet.
    ws = wb_macd.create_sheet(stock)
    
    #Writing Titles in Stock Sheet.
    ws.cell(1, 1, 'Date')
    ws.cell(1, 2, 'Open')
    ws.cell(1, 3, 'High')
    ws.cell(1, 4, 'Low')
    ws.cell(1, 5, 'Close')
    ws.cell(1, 6, 'Adj Close')
    ws.cell(1, 7, 'Volume')
    ws.cell(1, 8, 'MACD')
    ws.cell(1, 9, 'Signal Line')
    ws.cell(1, 10, 'Buy_Signal_Price')
    ws.cell(1, 11, 'Sell_Signal_Price')
    
    #Transferring data from dataframe to stock Excel sheet.
    for i in range(len(df)):
        cell = ws.cell(i+2, 1, str(df.loc[i, 'Date']))
        cell = ws.cell(i+2, 2, float(df.loc[i, 'Open']))
        cell = ws.cell(i+2, 3, float(df.loc[i, 'High']))
        cell = ws.cell(i+2, 4, float(df.loc[i, 'Low']))
        cell = ws.cell(i+2, 5, float(df.loc[i, 'Close']))
        cell = ws.cell(i+2, 6, float(df.loc[i, 'Adj Close']))
        cell = ws.cell(i+2, 7, float(df.loc[i, 'Volume']))        
        cell = ws.cell(i+2, 8, float(df.loc[i, 'MACD']))
        cell = ws.cell(i+2, 9, float(df.loc[i, 'Signal Line']))
        cell = ws.cell(i+2, 10, float(df.loc[i, 'Buy_Signal_Price']))
        cell = ws.cell(i+2, 11, float(df.loc[i, 'Sell_Signal_Price']))
    
    #Alignment for all Cells.
    cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
    for row in range(1, len(df)+2):
        for col in cols:
            cell = ws['{}{}'.format(col, row)]
            cell.font = font
            cell.border = border
            if col == 'A':
                cell.alignment = left_alignment
            else:
                cell.alignment = right_alignment
                cell.number_format = fmt_flo
    
    #Re-Alignment for title cells.
    for col in range(1, 12):
        cell = ws.cell(1, col)
        cell.font = font
        cell.border = border
        cell.alignment = center_alignment
    
    #Adding HyperLinks for plots.
    #1.Stock Close Price Plot.
    cell = ws.cell(2, 13, stkplt_hlink)
    cell.alignment = left_alignment
    cell.font = font
    cell.border = border
    #2.MACD Chart.
    cell = ws.cell(2, 14, hlinks[0])
    cell.alignment = left_alignment
    cell.font = font
    cell.border = border
    #3.MACD Signal Chart.
    cell = ws.cell(2, 15, hlinks[1])
    cell.alignment = left_alignment
    cell.font = font
    cell.border = border
    
    #Updating Cell Widths.
    dim_holder = DimensionHolder(worksheet=ws)
    for col in range(ws.min_column, ws.max_column+1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=12)
    ws.column_dimensions = dim_holder
    
    return ws


def macd_summary(stock, df, wb_macd, left_alignment, center_alignment, 
                    right_alignment, font, border, fmt_flo, main_row, 
                    last_entry):
    #Creating a summary MACD sheet with all Stock values.
    ws_main = wb_macd.active
    #Adding Titles
    ws_main.cell(1, 1, 'Stocks')
    ws_main.cell(1, 2, 'Last MACD Value')
    ws_main.cell(1, 3, 'Cur. Stock Price')
    ws_main.cell(1, 4, 'Price Diff. in %')
    ws_main.cell(1, 5, 'Remarks')
    for col in range(1, 6):
        cell = ws_main.cell(1, col)
        cell.font = font
        cell.border = border
        cell.alignment = center_alignment
        
    #Adding Stock name.
    cell = ws_main.cell(main_row, 1, stock)
    cell.alignment = left_alignment
    cell.font = font
    cell.border = border
    
    #Initializing variables.
    buy_location = 0
    sell_location = 0
    cell = ws_main.cell(1, 1)
    diff = 0
    actual_signal = ""
    days = 0
    percent_decimals = 0
    percent = 0
    
    #Naming fill patterns.
    lred_fill = PatternFill(start_color="FFB3B3",
                            end_color="FFB3B3",
                            fill_type="solid")
    lgreen_fill = PatternFill(start_color="85FFB1",
                              end_color="85FFB1",
                              fill_type="solid")
    lgrey_fill = PatternFill(start_color="B3B3B3",
                              end_color="B3B3B3",
                              fill_type="solid")
    
    #Adding Last Signal Value.
    for i in range(len(df)):
        if float(df.loc[i, 'Adj Close']) - float(df.loc[i, 'Buy_Signal_Price']) == 0 or float(df.loc[i, 'Close']) - float(df.loc[i, 'Buy_Signal_Price']) == 0:
            buy_location = i
        if float(df.loc[i, 'Adj Close']) - float(df.loc[i, 'Sell_Signal_Price']) == 0 or float(df.loc[i, 'Close']) - float(df.loc[i, 'Sell_Signal_Price']) == 0:
            sell_location = i
    if buy_location > sell_location:
        cell = ws_main.cell(main_row, 2, float(df.loc[buy_location, 'Buy_Signal_Price']))
        diff = float(df['Adj Close'].iloc[-1]) - float(df.loc[buy_location, 'Buy_Signal_Price'])
        actual_signal = 'Buy Signal'
        days = len(df) - buy_location
        cell.fill = lgreen_fill
    if sell_location > buy_location:
        cell = ws_main.cell(main_row, 2, float(df.loc[sell_location, 'Sell_Signal_Price']))
        diff = float(df['Adj Close'].iloc[-1]) - float(df.loc[sell_location, 'Sell_Signal_Price'])
        actual_signal = 'Sell Signal'
        days = len(df) - sell_location
        cell.fill = lred_fill
    cell.alignment = right_alignment
    cell.number_format = fmt_flo
    cell.font = font
    cell.border = border
    
    #CLOSE PRICE
    cell = ws_main.cell(main_row, 3, float(df['Adj Close'].iloc[-1]))
    if (float(df.loc[buy_location, 'Buy_Signal_Price']) - float(df['Adj Close'].iloc[-1])) < 0:
        cell.fill = lgreen_fill
    elif (float(df.loc[buy_location, 'Buy_Signal_Price']) - float(df['Adj Close'].iloc[-1])) > 0:
        cell.fill = lred_fill
    else:
        cell.fill = lgrey_fill
    cell.alignment = right_alignment
    cell.number_format = fmt_flo
    cell.font = font
    cell.border = border
    
    #FINDING PERCENTAGE.
    percent_decimals = (diff / float(df['Adj Close'].iloc[-1]))
    percent = percent_decimals * 100
    cell = ws_main.cell(main_row, 4, float(percent_decimals))
    if percent < 0:
        cell.fill = lred_fill
    elif percent > 0:
        cell.fill = lgreen_fill
    else:
        cell.fill = lgrey_fill
    cell.alignment = center_alignment
    fmt_flo_percent = u'0.00%'
    cell.number_format = fmt_flo_percent
    cell.font = font
    cell.border = border
    
    #REMARKS
    remarks = '{}, {} days before.'.format(actual_signal, days)
    cell = ws_main.cell(main_row, 5, remarks)
    cell.alignment = left_alignment
    cell.font = font
    cell.border = border
    
    if last_entry:
        #Updating First Sheet Name.
        ws_main.title = 'MACD Summary'
        
        #Updating Cell Widths.
        dim_holder = DimensionHolder(worksheet=ws_main)
        for col in range(ws_main.min_column, ws_main.max_column+1):
            dim_holder[get_column_letter(col)] = ColumnDimension(ws_main, min=col, max=col, width=12)
        ws_main.column_dimensions = dim_holder
        
        last_entry = False
                                
    
def buy_sell(signal):
    #Creating Buy and Sell Signals.
    Buy = []
    Sell = []
    flag = -1
    
    for i in range(0, len(signal)):
        if signal['MACD'][i] > signal['Signal Line'][i]:
            Sell.append(np.nan)
            if flag != 1:
                Buy.append(signal['Close'][i])
                flag = 1
            else:
                Buy.append(np.nan)
        elif signal['MACD'][i] < signal['Signal Line'][i]:
            Buy.append(np.nan)
            if flag != 0:
                Sell.append(signal['Close'][i])
                flag = 0
            else:
                Sell.append(np.nan)
        else:
            Buy.append(np.nan)
            Sell.append(np.nan)
    
    return(Buy, Sell)

def MACD(stock, df, wb_macd, left_alignment, center_alignment, 
            right_alignment, font, border, fmt_flo, stkplt_hlink, 
            plot_full_location, main_row, last_entry):
    #Calculate the MACF and signal line indicators.
    #Calculate the Short Term Exponential Moving Average (EMA)
    ShortEMA = df.Close.ewm(span=12, adjust=False).mean()
    #Calculate the Long Term Exponential Moving Average (EMA)
    LongEMA = df.Close.ewm(span=26, adjust=False).mean()
    #Calsulate the MACD line.
    MACD = ShortEMA - LongEMA
    #Calsulate the signal line.
    signal = MACD.ewm(span=9, adjust=False).mean()

    df['MACD'] = MACD
    df['Signal Line'] = signal
    
    a = buy_sell(df)
    df['Buy_Signal_Price'] = a[0]
    df['Sell_Signal_Price'] = a[1]
    
    hlinks = plot_macd(stock, df, MACD, plot_full_location)
    
    ws = stock_sheet(stock, df, wb_macd, left_alignment, center_alignment, 
                        right_alignment, font, border, fmt_flo, stkplt_hlink, 
                        plot_full_location, main_row, last_entry, hlinks)
    
    macd_summary(stock, df, wb_macd, left_alignment, center_alignment, 
                    right_alignment, font, border, fmt_flo, main_row, 
                    last_entry)
    
