import pandas as pd
import numpy as np
import openpyxl as xl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Border, Side, Alignment, Font
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
plt.style.use('fivethirtyeight')


def plot_dmac(stock, df, sma30, sma100, data, plot_full_location):
    #Plotting DMAC lines.
    plt.figure(figsize= (12.5, 4.5))
    plt.plot(df['Adj Close'], label=stock)
    plt.plot(sma30['Adj Close Price'], label='SMA30')
    plt.plot(sma100['Adj Close Price'], label='SMA100')
    plt.title('{} Adj. Close Price History'.format(stock))
    plt.xlabel('{} to {}'.format(df['Date'][0], df['Date'].iloc[-1]))
    plt.ylabel('Adj Close Price INR')
    plt.legend(loc='upper left')
    plt.savefig('{}/{}_DMAC.png'.format(plot_full_location, stock), dpi=150)
    dmac_hlink = '=HYPERLINK("{}\{}_DMAC.png","DMAC")'.format(plot_full_location, stock)
    plt.close()
    
    #Plotting Buy and Sell Signals
    plt.figure(figsize=(12.6, 4.6))
    plt.plot(data['Close'], label=stock, alpha=0.35)
    plt.plot(data['SMA30'], label='SMA30', alpha=0.35)
    plt.plot(data['SMA100'], label='SMA100', alpha=0.35)
    plt.scatter(data.index, data['Buy_Signal_Price'], label='Buy', marker='^', color='green')
    plt.scatter(data.index, data['Sell_Signal_Price'], label='Sell', marker='v', color='red')
    plt.title('{} Adj Close Price History Buy & Sell Signals'.format(stock))
    plt.xlabel('{} to {}'.format(df['Date'][0], df['Date'].iloc[-1]))
    plt.ylabel('Adj Close Price INR')
    plt.legend(loc='upper left')
    plt.savefig('{}/{}_DMAC_Signal.png'.format(plot_full_location, stock), dpi=150)
    signal_hlink = '=HYPERLINK("{}\{}_DMAC_Signal.png","DMAC SIGNAL")'.format(plot_full_location, stock)
    plt.close()
    
    return(dmac_hlink, signal_hlink)
    

def stock_sheet(stock, df, sma30, sma100, data, wb_dmac, left_alignment, center_alignment, 
                right_alignment, font, border, fmt_flo, stkplt_hlink_lt, 
                plot_full_location, main_row, last_entry, hlinks):
    #Transfer data from dataframes to Individual Stock Excel sheet.
    ws = wb_dmac.create_sheet(stock)
    
    #Writing Titles in Stock Sheet.
    ws.cell(1, 1, 'Date')
    ws.cell(1, 2, 'Open')
    ws.cell(1, 3, 'High')
    ws.cell(1, 4, 'Low')
    ws.cell(1, 5, 'Close')
    ws.cell(1, 6, 'Adj Close')
    ws.cell(1, 7, 'Volume')
    ws.cell(1, 8, 'SMA30')
    ws.cell(1, 9, 'SMA100')
    ws.cell(1, 10, 'Buy_Signal_Price')
    ws.cell(1, 11, 'Sell_Signal_Price')
    
    #Transferring data from dataframe to stock Excel sheet.
    for i in range(len(df)):
        cell = ws.cell(i+2, 1, str(df.loc[i, 'Date']))
        cell = ws.cell(i+2, 2, float(df.loc[i, 'Open']))
        cell = ws.cell(i+2, 3, float(df.loc[i, 'High']))
        cell = ws.cell(i+2, 4, float(df.loc[i, 'Low']))
        cell = ws.cell(i+2, 5, float(df.loc[i, 'Close']))
        cell = ws.cell(i+2, 6, float(df.loc[i, 'Adj Close']))
        cell = ws.cell(i+2, 7, float(df.loc[i, 'Volume']))
        cell = ws.cell(i+2, 8, float(data.loc[i, 'SMA30']))
        cell = ws.cell(i+2, 9, float(data.loc[i, 'SMA100']))
        cell = ws.cell(i+2, 10, float(data.loc[i, 'Buy_Signal_Price']))
        cell = ws.cell(i+2, 11, float(data.loc[i, 'Sell_Signal_Price']))
    
    #Alignment for all Cells.
    cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
    for row in range(1, len(df)+2):
        for col in cols:
            cell = ws['{}{}'.format(col, row)]
            cell.font = font
            cell.border = border
            if col == 'A':
                cell.alignment = left_alignment
            else:
                cell.alignment = right_alignment
                cell.number_format = fmt_flo
    
    #Re-Alignment for title cells.
    for col in range(1, 12):
        cell = ws.cell(1, col)
        cell.font = font
        cell.border = border
        cell.alignment = center_alignment
    
    #Adding HyperLinks for plots.
    #1.Stock Close Price Plot.
    cell = ws.cell(2, 13, stkplt_hlink_lt)
    cell.alignment = left_alignment
    cell.font = font
    cell.border = border
    #2.MACD Chart.
    cell = ws.cell(2, 14, hlinks[0])
    cell.alignment = left_alignment
    cell.font = font
    cell.border = border
    #3.MACD Signal Chart.
    cell = ws.cell(2, 15, hlinks[1])
    cell.alignment = left_alignment
    cell.font = font
    cell.border = border
    
    #Updating Cell Widths.
    dim_holder = DimensionHolder(worksheet=ws)
    for col in range(ws.min_column, ws.max_column+1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=12)
    ws.column_dimensions = dim_holder
    
    return ws


def dmac_summary(stock, df, sma30, sma100, data, wb_dmac, left_alignment, center_alignment, 
                    right_alignment, font, border, fmt_flo, main_row, 
                    last_entry):
    #Creating a summary DMAC sheet with all Stock values.
    ws_main = wb_dmac.active
    #Adding Titles
    ws_main.cell(1, 1, 'Stocks')
    ws_main.cell(1, 2, 'Last Crossover Value')
    ws_main.cell(1, 3, 'Cur. Stock Price')
    ws_main.cell(1, 4, 'Price Diff. in %')
    ws_main.cell(1, 5, 'Remarks')
    for col in range(1, 6):
        cell = ws_main.cell(1, col)
        cell.font = font
        cell.border = border
        cell.alignment = center_alignment
    
    #Adding Stock name.
    cell = ws_main.cell(main_row, 1, stock)
    cell.alignment = left_alignment
    cell.font = font
    cell.border = border
    
    #Naming fill patterns.
    lred_fill = PatternFill(start_color="FFB3B3",
                            end_color="FFB3B3",
                            fill_type="solid")
    lgreen_fill = PatternFill(start_color="85FFB1",
                              end_color="85FFB1",
                              fill_type="solid")
    lgrey_fill = PatternFill(start_color="B3B3B3",
                              end_color="B3B3B3",
                              fill_type="solid")
    
    #Adding last Crossover value.
    buy_location = 0
    sell_location = 0
    for i in reversed(range(len(df))):
        if buy_location == 0 & sell_location == 0:
            if float(df.loc[i, 'Adj Close']) - float(data.loc[i, 'Buy_Signal_Price']) == 0 or float(df.loc[i, 'Close']) - float(data.loc[i, 'Buy_Signal_Price']) == 0:
                buy_location = i
            if float(df.loc[i, 'Adj Close']) - float(data.loc[i, 'Sell_Signal_Price']) == 0 or float(df.loc[i, 'Close']) - float(data.loc[i, 'Sell_Signal_Price']) == 0:
                sell_location = i
        else:
            break
    if buy_location > sell_location:
        cell = ws_main.cell(main_row, 2, float(data.loc[buy_location, 'Buy_Signal_Price']))
        diff = float(df['Adj Close'].iloc[-1]) - float(data.loc[buy_location, 'Buy_Signal_Price'])
        actual_signal = 'Buy Signal'
        days = len(df) - buy_location
        cell.fill = lgreen_fill
    if sell_location > buy_location:
        cell = ws_main.cell(main_row, 2, float(data.loc[sell_location, 'Sell_Signal_Price']))
        diff = float(df['Adj Close'].iloc[-1]) - float(data.loc[sell_location, 'Sell_Signal_Price'])
        actual_signal = 'Sell Signal'
        days = len(df) - sell_location
        cell.fill = lred_fill
    cell.alignment = right_alignment
    cell.number_format = fmt_flo
    cell.font = font
    cell.border = border
    
    #CLOSE PRICE
    cell = ws_main.cell(main_row, 3, float(df['Adj Close'].iloc[-1]))
    if (float(data.loc[buy_location, 'Buy_Signal_Price']) - float(df['Adj Close'].iloc[-1])) < 0:
        cell.fill = lgreen_fill
    elif (float(data.loc[buy_location, 'Buy_Signal_Price']) - float(df['Adj Close'].iloc[-1])) > 0:
        cell.fill = lred_fill
    else:
        cell.fill = lgrey_fill
    cell.alignment = right_alignment
    cell.number_format = fmt_flo
    cell.font = font
    cell.border = border
    
     #FINDING PERCENTAGE.
    percent_decimals = (diff / float(df['Adj Close'].iloc[-1]))
    percent = percent_decimals * 100
    cell = ws_main.cell(main_row, 4, float(percent_decimals))
    if percent < 0:
        cell.fill = lred_fill
    elif percent > 0:
        cell.fill = lgreen_fill
    else:
        cell.fill = lgrey_fill
    cell.alignment = center_alignment
    fmt_flo_percent = u'0.00%'
    cell.number_format = fmt_flo_percent
    cell.font = font
    cell.border = border
    
    #REMARKS
    remarks = '{}, {} days before.'.format(actual_signal, days)
    cell = ws_main.cell(main_row, 5, remarks)
    cell.alignment = left_alignment
    cell.font = font
    cell.border = border
    
    if last_entry:
        #Updating First Sheet Name.
        ws_main.title = 'DMAC Summary'
        
        #Updating Cell Widths.
        dim_holder = DimensionHolder(worksheet=ws_main)
        for col in range(ws_main.min_column, ws_main.max_column+1):
            dim_holder[get_column_letter(col)] = ColumnDimension(ws_main, min=col, max=col, width=12)
        ws_main.column_dimensions = dim_holder
        
        last_entry = False


def buy_sell_dmac(data):
    sigPriceBuy = []
    sigPriceSell = []
    flag = -1

    for i in range(len(data)):
        if data['SMA30'][i] > data['SMA100'][i]:
            if flag != 1:
                sigPriceBuy.append(data['Close'][i])
                sigPriceSell.append(np.nan)
                flag = 1
            else:
                sigPriceBuy.append(np.nan)
                sigPriceSell.append(np.nan)
        elif data['SMA30'][i] < data['SMA100'][i]:
            if flag != 0:
                sigPriceBuy.append(np.nan)
                sigPriceSell.append(data['Close'][i])
                flag = 0
            else:
                sigPriceBuy.append(np.nan)
                sigPriceSell.append(np.nan)
        else:
            sigPriceBuy.append(np.nan)
            sigPriceSell.append(np.nan)
    return (sigPriceBuy, sigPriceSell)


def DMAC(stock, df, wb_dmac, left_alignment, center_alignment, 
            right_alignment, font, border, fmt_flo, stkplt_hlink_lt, 
            plot_full_location, main_row, last_entry):
    sma30 = pd.DataFrame()
    sma30['Adj Close Price'] = df['Adj Close'].rolling(window=30).mean()
    
    sma100 = pd.DataFrame()
    sma100['Adj Close Price'] = df['Adj Close'].rolling(window=100).mean()
    
    data = pd.DataFrame()
    data['Close'] = df['Adj Close']
    data['SMA30'] = sma30['Adj Close Price']
    data['SMA100'] = sma100['Adj Close Price']
    
    buy_sell = buy_sell_dmac(data)
    data['Buy_Signal_Price'] = buy_sell[0]
    data['Sell_Signal_Price'] = buy_sell[1]
    
    hlinks = plot_dmac(stock, df, sma30, sma100, data, plot_full_location)
    
    ws = stock_sheet(stock, df, sma30, sma100, data, wb_dmac, left_alignment, center_alignment, 
                        right_alignment, font, border, fmt_flo, stkplt_hlink_lt, 
                        plot_full_location, main_row, last_entry, hlinks)
    
    dmac_summary(stock, df, sma30, sma100, data, wb_dmac, left_alignment, center_alignment, 
                    right_alignment, font, border, fmt_flo, main_row, 
                    last_entry)
