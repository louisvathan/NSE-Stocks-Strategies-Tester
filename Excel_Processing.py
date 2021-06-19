import pandas as pd
import openpyxl as xl
from openpyxl.styles import Color, PatternFill, Border, Side, Alignment, Font
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter

fmt_flo = u'#,##0.00;-#,##0.00'
fmt_flo_percent = u'#,##0.00%;-#,##0.00%'
fmt_flo_rupees = u'Rs.#,##0.00;-Rs.#,##0.00'
center_alignment = Alignment(horizontal='center',
                             vertical='center',
                             wrap_text=True)
left_alignment = Alignment(horizontal='left',
                           vertical='center',
                           wrap_text=True)
right_alignment = Alignment(horizontal='right',
                            vertical='center',
                            wrap_text=True)
font = Font(name='Arial',
            size=8,
            color='FF000000')
border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))
lred_fill = PatternFill(start_color="FFB3B3",
                        end_color="FFB3B3",
                        fill_type="solid")
lgreen_fill = PatternFill(start_color="85FFB1",
                            end_color="85FFB1",
                            fill_type="solid")
lgrey_fill = PatternFill(start_color="B3B3B3",
                            end_color="B3B3B3",
                            fill_type="solid")

def CellWidths(sheet, width):
    #Updating Cell Widths.
    dim_holder = DimensionHolder(worksheet=sheet)
    for col in range(sheet.min_column, sheet.max_column+1):
        dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=width)
    sheet.column_dimensions = dim_holder


def excel_conv(indicator, df, broker_df, value_df, result_df, wb):
    ws = wb.create_sheet(indicator)

    df['Cash'] = value_df['Cash']
    df['Holding_Value'] = broker_df['Value']
    df['Holding'] = broker_df['Holding']
    df['P&L'] = result_df['P&L']
    df['Portfolio_Value'] = value_df['Portfolio_Value']

    titles = df.columns.tolist()

    for i in range(len(titles)):
        cell = ws.cell(1, i+1, titles[i])
        cell.font = font
        cell.border = border
        cell.alignment = center_alignment
        for j in range(len(df)):
            if i == 0:
                cell = ws.cell(j+2, i+1, str(df.loc[j, titles[i]]))
                cell.alignment = left_alignment
                cell.font = font
                cell.border = border
            else:
                cell = ws.cell(j+2, i+1, float(df.loc[j, titles[i]]))
                cell.alignment = right_alignment
                cell.number_format = fmt_flo
                cell.font = font
                cell.border = border

    CellWidths(ws, 12)


def stock_summary_sheet(stock, wb, stock_summary, index, startegy_descp, stkplt_hlink):
    ws_main = wb['Sheet']
    cell = ws_main.cell(1, 1, 'Strategies')
    cell.font = font
    cell.border = border
    cell.alignment = center_alignment
    
    for i in range(len(index)):
        cell = ws_main.cell(i+2, 1, index[i])
        cell.font = font
        cell.border = border
        cell.alignment = left_alignment
        for j in range(len(startegy_descp)):
            cell = ws_main.cell(1, j+2, startegy_descp[j])
            cell.font = font
            cell.border = border
            cell.alignment = center_alignment

            text_rows = ['Current Stock Details', 'Last Signal', 'Remarks', 'Backtest Results', 'Starting Date', 'Ending Date',
                         'Backtest Plot', 'Stock Plot', 'Strategy Plot', 'Signals Plot']
            percent_rows = ['Price Difference(%)', 'Percent used to trade(%)', 'Profit & Loss(%)', 'Positive Trade Percentage(%)']
            
            if index[i] in text_rows:
                cell = ws_main.cell(i+2, j+2, str(stock_summary.loc[index[i], startegy_descp[j]]))
            elif index[i] in percent_rows:
                cell = ws_main.cell(i+2, j+2, float(stock_summary.loc[index[i], startegy_descp[j]]))
                cell.number_format = fmt_flo_percent
            else:
                cell = ws_main.cell(i+2, j+2, float(stock_summary.loc[index[i], startegy_descp[j]]))
                cell.number_format = fmt_flo
            cell.font = font
            cell.border = border
            cell.alignment = center_alignment

            cell = ws_main.cell(i+2, j+2)
            if cell.value == 'Buy':
                cell.fill = lgreen_fill
                ws_main.cell(i+3, j+2).fill = lgreen_fill
                if float(stock_summary.loc[index[i+1], startegy_descp[j]]) < float(stock_summary.loc[index[i+2], startegy_descp[j]]):
                    ws_main.cell(i+4, j+2).fill = lgreen_fill
                    ws_main.cell(i+5, j+2).fill = lgreen_fill
                elif float(stock_summary.loc[index[i+1], startegy_descp[j]]) > float(stock_summary.loc[index[i+2], startegy_descp[j]]):
                    ws_main.cell(i+4, j+2).fill = lred_fill
                    ws_main.cell(i+5, j+2).fill = lred_fill
                ws_main.cell(i+6, j+2).fill = lgreen_fill
            elif cell.value == 'Sell':
                cell.fill = lred_fill
                ws_main.cell(i+3, j+2).fill = lred_fill
                if float(stock_summary.loc[index[i+1], startegy_descp[j]]) > float(stock_summary.loc[index[i+2], startegy_descp[j]]):
                    ws_main.cell(i+4, j+2).fill = lred_fill
                    ws_main.cell(i+5, j+2).fill = lred_fill
                elif float(stock_summary.loc[index[i+1], startegy_descp[j]]) < float(stock_summary.loc[index[i+2], startegy_descp[j]]):
                    ws_main.cell(i+4, j+2).fill = lgreen_fill
                    ws_main.cell(i+5, j+2).fill = lgreen_fill
                ws_main.cell(i+6, j+2).fill = lred_fill


    CellWidths(ws_main, 20)
    ws_main.title = '{} Summary'.format(stock)


def full_summary(wb_summary, summary, date, folder_full_location):
    ws_summary = wb_summary.active

    cell = ws_summary.cell(1, 1, 'Strategies')
    cell.font = font
    cell.border = border
    cell.alignment = center_alignment

    titles = summary.columns.tolist()
    index = summary.index.values.tolist()

    for i in range(len(index)):
        stock_filename = '{}.xlsx'.format(index[i])
        stock_file = folder_full_location + '/' + stock_filename
        stock_hlink = '=HYPERLINK("{}","{}")'.format(stock_file, index[i])
        cell = ws_summary.cell(i+2, 1, stock_hlink)
        cell.font = font
        cell.border = border
        cell.alignment = left_alignment
        for j in range(len(titles)):
            cell = ws_summary.cell(1, j+2, titles[j])
            cell.font = font
            cell.border = border
            cell.alignment = center_alignment

            cell = ws_summary.cell(i+2, j+2, str(summary.loc[index[i], titles[j]]))
            cell.font = font
            cell.border = border
            cell.alignment = center_alignment
            words = cell.value.split()
            if words[0] == 'Buy':
                cell.fill = lgreen_fill
            elif words[0] == 'Sell':
                cell.fill = lred_fill
            
    CellWidths(ws_summary, 20)
    ws_summary.title = 'Summary ({})'.format(date)
