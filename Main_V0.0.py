import pandas as pd
#import numpy as np
import os
import time
import matplotlib.pyplot as plt
import openpyxl as xl
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Border, Side, Alignment, Font
#from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
#from openpyxl.styles.differential import DifferentialStyle
#from openpyxl.formatting import Rule
#from openpyxl.formatting.rule import ColorScaleRule
#from openpyxl.utils import get_column_letter
import datetime
import requests
plt.style.use('fivethirtyeight')

#Importing Functions from other files.
from RSI import *
from MACD import *
from DMAC import *

#Giving Imputs
nifty = ['SBIN', 'TATASTEEL', 'DIVISLAB', 'KOTAKBANK', 'DRREDDY', 'ULTRACEMCO',
         'ITC', 'CIPLA', 'HDFCBANK', 'HINDUNILVR', 'SUNPHARMA', 'JSWSTEEL', 'NTPC',
         'HDFC', 'POWERGRID', 'BAJFINANCE', 'HINDALCO', 'NESTLEIND', 'ONGC', 'HDFCLIFE',
         'SBILIFE', 'BAJAJFINSV', 'RELIANCE', 'TITAN', 'BAJAJ-AUTO', 'ADANIPORTS',
         'GRASIM', 'INDUSINDBK', 'M&M', 'HEROMOTOCO', 'INFY', 'BPCL', 'WIPRO',
         'ASIANPAINT', 'TCS', 'LT', 'TECHM', 'HCLTECH', 'BRITANNIA', 'IOC', 'EICHERMOT',
         'SHREECEM', 'GAIL', 'ICICIBANK', 'MARUTI', 'UPL', 'COALINDIA', 'TATAMOTORS',
         'BHARTIARTL', 'AXISBANK']

nifty_below2000 = ['IOC', 'NTPC', 'ONGC', 'COALINDIA', 'GAIL', 'ITC', 'POWERGRID',
                   'HINDALCO', 'TATAMOTORS', 'JSWSTEEL', 'BPCL', 'SBIN', 'WIPRO',
                   'UPL', 'BHARTIARTL', 'SUNPHARMA', 'ICICIBANK', 'ADANIPORTS',
                   'TATASTEEL', 'HDFCLIFE', 'AXISBANK', 'CIPLA', 'SBILIFE', 'M&M',
                   'HCLTECH', 'TECHM', 'INDUSINDBK', 'GRASIM', 'INFY', 'TITAN',
                   'LT', 'HDFCBANK', 'KOTAKBANK', 'RELIANCE', 'HINDUNILVR']

nifty_below1000 = ['IOC', 'NTPC', 'ONGC', 'COALINDIA', 'GAIL', 'ITC', 'POWERGRID',
                   'HINDALCO', 'TATAMOTORS', 'JSWSTEEL', 'BPCL', 'SBIN', 'WIPRO',
                   'UPL', 'BHARTIARTL', 'SUNPHARMA', 'ICICIBANK', 'ADANIPORTS',
                   'TATASTEEL', 'HDFCLIFE', 'AXISBANK', 'CIPLA', 'SBILIFE', 'M&M',
                   'HCLTECH', 'TECHM', 'INDUSINDBK', 'GRASIM']

test_stocks = ['SBIN', 'TATASTEEL', 'DIVISLAB', 'KOTAKBANK', 'DRREDDY', 'ULTRACEMCO',
               'ITC', 'CIPLA', 'HDFCBANK', 'HINDUNILVR', 'SUNPHARMA', 'JSWSTEEL', 'NTPC']

mystocks = ['HDFCBANK', 'SBIN']

stocks = nifty_below1000
period1 = '1457136000'
period2 = '1614902400'

#Plots the stock price in line chart.
def plot_stock(plt_df, plot_full_location):
    plt_df = plt_df.set_index(pd.DatetimeIndex(plt_df['Date'].values))
    
    plt.figure(figsize=(12.2, 4.5))
    plt.plot(plt_df.index, plt_df['Adj Close'], label='Adj Close Price')
    plt.title('{} Close Price History'.format(stock))
    plt.xlabel('Date', fontsize=18)
    plt.ylabel('Adj. Close Price in INR', fontsize=18)
    plt.savefig('{}/{}_plot({} to {}).png'.format(plot_full_location, stock, plt_df['Date'][0], plt_df['Date'].iloc[-1]), dpi=150)
    stkplt_hlink = '=HYPERLINK("{}\{}_plot({} to {}).png","PLOT")'.format(plot_full_location, stock, plt_df['Date'][0], plt_df['Date'].iloc[-1])
    plt.close()
    return stkplt_hlink
    
#Format Options for Excel Sheet.
fmt_num = u'0;'
fmt_flo = u'#,###0.000;-#,###0.000'
center_alignment = Alignment(horizontal='center',
                             vertical='center',
                             wrap_text=True)
left_alignment = Alignment(horizontal='left',
                           vertical='center',
                           wrap_text=True)
right_alignment = Alignment(horizontal='right',
                            vertical='center',
                            wrap_text=True)
font = Font(name='Arial',
            size=8,
            color='FF000000')
border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))

#Initializing Workbooks
wb_rsi = Workbook()
wb_macd = Workbook()
wb_dmac = Workbook()

now = datetime.datetime.now()
date = str(now.year) + '{:02d}'.format(now.month) + '{:02d}'.format(now.day)

#CREATING MAIN AND CSV FOLDERS.
current_location = os.getcwd()

foldername = 'Results(' + date + ')'
csv_location = foldername + '/CSV FILES'
plot_location = foldername + '/PLOTS'

try:
    os.mkdir(foldername)
except FileExistsError:
    pass

try:
    os.mkdir(csv_location)
except FileExistsError:
    pass

try:
    os.mkdir(plot_location)
except FileExistsError:
    pass

csv_full_location = current_location + '/' + csv_location
plot_full_location = current_location + '/' + plot_location

main_row = 2
last_entry = False
for stock in stocks:
    csv_name = '{}-{}-{}.csv'.format(stock, period1, period2)
    csv_file = csv_full_location + '/' + csv_name
    for root, dir, files in os.walk(csv_full_location):
        if csv_name in files:
            df_lt = pd.read_csv(csv_file)
            delay = 0
        else:
            csv_url = "https://query1.finance.yahoo.com/v7/finance/download/{}.NS?period1={}&period2={}&interval=1d&events=history".format(stock, period1, period2)
            r = requests.get(csv_url, allow_redirects=True)
            open(csv_file, 'wb').write(r.content)
            df_lt = pd.read_csv(csv_file)
            delay = 2
    
    df_lt = df_lt.dropna()
    df_lt = df_lt.reset_index()
    df_lt.drop(['index'], axis=1, inplace=True)
    
    stkplt_hlink_lt = plot_stock(df_lt, plot_full_location)
    
    df_st = df_lt.tail(150)
    df_st = df_st.reset_index()
    df_st.drop(['index'], axis=1, inplace=True)
    
    stkplt_hlink_st = plot_stock(df_st, plot_full_location)
    
    if stocks.index(stock)+1 == len(stocks):
        last_entry = True
    
    RSI(stock, df_st, wb_rsi, left_alignment, center_alignment, 
        right_alignment, font, border, fmt_flo, stkplt_hlink_st, 
        plot_full_location, main_row, last_entry)
    
    MACD(stock, df_st, wb_macd, left_alignment, center_alignment, 
            right_alignment, font, border, fmt_flo, stkplt_hlink_st, 
            plot_full_location, main_row, last_entry)
    
    DMAC(stock, df_lt, wb_dmac, left_alignment, center_alignment, 
            right_alignment, font, border, fmt_flo, stkplt_hlink_lt, 
            plot_full_location, main_row, last_entry)
    
    main_row += 1
    
    print('{} Done.'.format(stock))
    
    time.sleep(delay)

wb_rsi.save('{}/{}/RSI.xlsx'.format(current_location, foldername))
wb_macd.save('{}/{}/MACD.xlsx'.format(current_location, foldername))
wb_dmac.save('{}/{}/DMAC.xlsx'.format(current_location, foldername))
